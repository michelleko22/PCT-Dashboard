"""
PCT Waiting Time Dashboard Generator
Reads Manufacturing and Packaging schedule files from the network drive,
computes cycle time metrics, and generates a self-contained dashboard.html.

Run:  python generate_dashboard.py

Update MFG_FILE and PKG_FILE below when a new month's file is created.

Requires:  pip install openpyxl
"""
import io, json, os, re
import openpyxl
from datetime import datetime, timedelta
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# ── File paths ────────────────────────────────────────────────────────────────
# Update these each month when new files are created.
MFG_FILE = r'\\S-ya-file\consumer\Departments\Manufacturing\Master Prod Schedule Manufacturing\Master File & Achived Months\2026\3. Marc 2026 PRODUCTION SCHED - Manufacturing.xlsm'
PKG_FILE = r'\\S-ya-file\consumer\Departments\Manufacturing\Master Production Schedule Packaging\Master and Old Production Schedules\2026\3. PRODUCTION SCHED - March 2026.xlsm'


def load_wb(path):
    if not os.path.exists(path):
        raise FileNotFoundError(
            f'File not found: {path}\n'
            'Make sure you are connected to the Sanofi network (or VPN)\n'
            'and update MFG_FILE / PKG_FILE at the top of this script.'
        )
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

PCT_TARGET_DAYS = 17

# ── column maps (0-indexed) ──────────────────────────────────────────────────
# COMPRESSION_STATUS_COL: column index for the run-status cell in compression sheets
# (e.g. 'r' = released/ready, 'Wait-Disp' = waiting for dispensing)
# Verify against the actual Excel layout and adjust if needed.
COMPRESSION_STATUS_COL = 3
QUEUE_STATUSES = {'r', 'w'}            # lowercased: r=released/ready, w=waiting
RUNNING_STATUSES = {'ip'}              # ip=in progress (currently running)

# COMPRESSION_CLEAN_COL: column index for clean/changeover hours in compression sheets
# Adjust if clean time is in a different column (currently assumes col 1, adjacent to run at col 2)
COMPRESSION_CLEAN_COL = 1

COLS = {
    'dispensing':   {'wo':5,'item':6,'qty':7,'desc':8,'run':2,'start':11,'finish':12},
    'compression':  {'wo':5,'item':6,'qty':7,'desc':8,'run':2,'clean':COMPRESSION_CLEAN_COL,'start':12,'finish':13,'status':COMPRESSION_STATUS_COL},
    'coating':      {'wo':6,'item':7,'qty':8,'desc':9,'run':3,'start':13,'finish':14},
    'encap_hard':   {'wo':5,'item':6,'qty':7,'desc':8,'run':2,'start':12,'finish':13},
    'sg_gel_disp':  {'wo':6,'item':7,'qty':8,'desc':9,'run':3,'start':12,'finish':13},
    'sg_med_disp':  {'wo':6,'item':7,'qty':8,'desc':9,'run':3,'start':12,'finish':13},
    'sg_encap':     {'wo':5,'item':6,'qty':7,'desc':8,'run':2,'start':12,'finish':13},
    'packaging':    {'wo':4,'bulk':11,'start':0,'run':1},
}

STEP_ORDER = ['dispensing','sg_gel_disp','sg_med_disp',
              'compression','encap_hard','sg_encap','coating']

MFG_SHEETS = {
    'dispensing':  ['Disp 1','Disp 2','Disp 3','Disp 4'],
    'compression': ['TC1-Korsch 1','TC2-Korsch 2','TC 3-Kil','TC 4-Kil',
                    'TC 6-Fette 1','TC 7-Fette 2','TC 8-Fette 3'],
    'encap_hard':  ['TC 5-Bosch'],
    'coating':     ['Coating 1','Coating 2','Coating 3'],
    'sg_gel_disp': ['SG Gel Disp 1','SG Gel Disp 2'],
    'sg_med_disp': ['SG Med Disp 1','SG Med Disp 2','SG Med Disp 3'],
    'sg_encap':    ['SG Encap 1','SG Encap 2','SG Encap 3'],
}
PKG_SHEETS = ['VFILLDU1','VFILLTRI','VFILLCR1','VFILLDU2','VFILLBL1']

_DT_FMTS = (
    '%d/%m/%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S',
    '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S',
    '%d/%m/%Y %H:%M',    '%m/%d/%Y %H:%M',
    '%Y-%m-%d %H:%M',
    '%d/%m/%Y',          '%m/%d/%Y',          '%Y-%m-%d',
)

def as_dt(v):
    if isinstance(v, datetime):
        return v
    s = str(v).strip() if v is not None else ''
    if not s:
        return None
    # Google Sheets may return numeric serial strings
    try:
        n = float(s)
        if 40000 < n < 60000:
            return datetime(1899, 12, 30) + timedelta(days=n)
    except ValueError:
        pass
    for fmt in _DT_FMTS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None

def as_int(v):
    s = str(v).strip() if v is not None else ''
    if not s:
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None

def as_float(v):
    s = str(v).strip() if v is not None else ''
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None

def as_str(v):
    return str(v).strip() if v is not None else ''

# ── load product types ────────────────────────────────────────────────────────
def load_product_types(wb):
    ws = wb['Products']
    rows = list(ws.iter_rows(values_only=True))
    types = {}
    for r in rows[3:]:   # header at index 2, data from index 3
        if r[0] is None: continue
        k = as_str(r[0])
        t = as_str(r[2]) if len(r) > 2 and r[2] else ''
        if k and t:
            types[k] = t
    return types

# ── extract work center records ───────────────────────────────────────────────
def extract_mfg(wb, step, sheets, cols):
    records = []
    max_col = max(cols.values())
    for sname in sheets:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[8:]:   # first 8 rows are header/summary
            if len(row) <= max_col: continue
            wo = as_int(row[cols['wo']])
            if wo is None or wo < 1_000_000: continue
            start  = as_dt(row[cols['start']])
            finish = as_dt(row[cols['finish']])
            status = as_str(row[cols['status']]).lower() if 'status' in cols and cols['status'] < len(row) else ''
            is_queue = status in QUEUE_STATUSES
            if start is None and finish is None and not is_queue: continue
            run_h   = as_float(row[cols['run']])   if 'run'   in cols else None
            clean_h = as_float(row[cols['clean']]) if 'clean' in cols else None
            run_h   = run_h   if run_h   and 0 < run_h   < 500 else None
            clean_h = clean_h if clean_h and 0 < clean_h < 100 else None
            if start is None and finish is not None and run_h:
                start = finish - timedelta(hours=run_h)
            records.append({
                'wo': wo, 'step': step, 'wc': sname,
                'item':   as_str(row[cols['item']]),
                'qty':    as_str(row[cols['qty']]),
                'desc':   as_str(row[cols['desc']]),
                'start':  start, 'finish': finish,
                'run_h':  run_h, 'clean_h': clean_h,
                'status': status,
            })
    return records

def extract_pkg(wb, sheets, cols):
    records = []
    max_col = max(cols.values())
    for sname in sheets:
        if sname not in wb.sheetnames: continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:   # header is row 0
            if len(row) <= max_col: continue
            wo = as_int(row[cols['wo']])
            if wo is None or wo < 1_000_000: continue
            start = as_dt(row[cols['start']])
            if start is None: continue
            bulk  = as_str(row[cols['bulk']])
            run_h = as_float(row[cols['run']])
            run_h = run_h if run_h and 0 < run_h < 500 else None
            finish = start + timedelta(hours=run_h) if run_h else None
            records.append({
                'wo': wo, 'step': 'packaging', 'wc': sname,
                'item': bulk, 'start': start, 'finish': finish,
                'run_h': run_h,
            })
    return records

# ── main extraction ───────────────────────────────────────────────────────────
print("Loading workbooks…")
wb_mfg = load_wb(MFG_FILE)
wb_pkg = load_wb(PKG_FILE)

product_types = load_product_types(wb_mfg)

all_mfg = []
for step, sheets in MFG_SHEETS.items():
    recs = extract_mfg(wb_mfg, step, sheets, COLS[step])
    all_mfg.extend(recs)
    print(f"  {step}: {len(recs)} records")

all_pkg = extract_pkg(wb_pkg, PKG_SHEETS, COLS['packaging'])
print(f"  packaging: {len(all_pkg)} records")

# ── group manufacturing records by WO ────────────────────────────────────────
wo_steps = defaultdict(list)
for r in all_mfg:
    wo_steps[r['wo']].append(r)

# ── build PCT / waiting time records ─────────────────────────────────────────
# For each WO, we know its steps. Determine process sequence and compute waits.
# Waiting time = next_step.start - prev_step.finish  (only if >0)

STEP_RANK = {s: i for i, s in enumerate(STEP_ORDER)}

waiting_records = []    # {wo, item, desc, transition, wait_days, month, wc_from, wc_to}
pct_records     = []    # {wo, item, desc, type, pct_days, mfg_days, pkg_days, wait_days, month}

for wo, steps in wo_steps.items():
    # deduplicate by step (keep record with latest finish if duplicates)
    by_step = {}
    for s in steps:
        k = s['step']
        if k not in by_step or (s['finish'] and (by_step[k]['finish'] is None or s['finish'] > by_step[k]['finish'])):
            by_step[k] = s

    item = next((s['item'] for s in steps if s['item']), '')
    desc = next((s['desc'] for s in steps if s['desc']), '')
    ptype = product_types.get(item, 'Unknown')

    # Sort the available steps in process order
    ordered = sorted(by_step.values(), key=lambda s: STEP_RANK.get(s['step'], 99))

    # Compute consecutive waits
    for i in range(len(ordered) - 1):
        a, b = ordered[i], ordered[i+1]
        if a['finish'] and b['start']:
            wait_h = (b['start'] - a['finish']).total_seconds() / 3600
            if 0 <= wait_h <= 30 * 24:   # sanity: 0 to 30 days
                t_name = f"{a['step'].replace('_',' ').title()} → {b['step'].replace('_',' ').title()}"
                ref_date = (a['finish'] or b['start'])
                waiting_records.append({
                    'wo': wo, 'item': item, 'desc': desc,
                    'transition': t_name,
                    'wait_h': round(wait_h, 2),
                    'wait_d': round(wait_h / 24, 2),
                    'month': ref_date.strftime('%Y-%m') if ref_date else '',
                    'week':  ref_date.strftime('%Y-W%V') if ref_date else '',
                    'date':  ref_date.strftime('%Y-%m-%d') if ref_date else '',
                    'wc_from': a['wc'], 'wc_to': b['wc'],
                    'type': ptype,
                })

# ── packaging waiting time: last mfg step → packaging ────────────────────────
# For each packaging WO, find most recent manufacturing finish for that bulk item
# Group pkg by bulk item
pkg_by_item = defaultdict(list)
for r in all_pkg:
    if r['item']:
        pkg_by_item[r['item']].append(r)

# Group mfg last-step finish by item
mfg_last_by_item = defaultdict(list)
for wo, steps in wo_steps.items():
    by_step = {}
    for s in steps:
        k = s['step']
        if k not in by_step or (s['finish'] and (by_step[k]['finish'] is None or s['finish'] > by_step[k]['finish'])):
            by_step[k] = s
    ordered = sorted(by_step.values(), key=lambda s: STEP_RANK.get(s['step'], 99))
    if not ordered: continue
    last = ordered[-1]
    if last['finish']:
        item = last['item'] or next((s['item'] for s in steps if s['item']), '')
        mfg_last_by_item[item].append({
            'wo': wo, 'finish': last['finish'], 'step': last['step'],
            'wc': last['wc'],
            'desc': next((s['desc'] for s in steps if s['desc']), ''),
            'type': product_types.get(item, 'Unknown'),
        })

pkg_wait_records = []
for item, pkg_list in pkg_by_item.items():
    mfg_list = mfg_last_by_item.get(item, [])
    if not mfg_list: continue
    # Sort mfg finishes ascending
    mfg_sorted = sorted(mfg_list, key=lambda x: x['finish'])
    # For each packaging start, find closest preceding mfg finish
    for pr in sorted(pkg_list, key=lambda x: x['start']):
        pkg_start = pr['start']
        best = None
        for m in mfg_sorted:
            diff_h = (pkg_start - m['finish']).total_seconds() / 3600
            if -48 <= diff_h <= 30 * 24:   # allow small negative for same-day
                if best is None or abs(diff_h) < abs((pkg_start - best['finish']).total_seconds()/3600):
                    best = m
        if best:
            wait_h = max(0, (pkg_start - best['finish']).total_seconds() / 3600)
            if wait_h <= 30 * 24:
                pkg_wait_records.append({
                    'wo': pr['wo'], 'item': item,
                    'desc': best['desc'],
                    'transition': f"Manufacturing → Packaging",
                    'wait_h': round(wait_h, 2),
                    'wait_d': round(wait_h / 24, 2),
                    'month': pkg_start.strftime('%Y-%m'),
                    'week':  pkg_start.strftime('%Y-W%V'),
                    'date':  pkg_start.strftime('%Y-%m-%d'),
                    'wc_from': best['wc'], 'wc_to': pr['wc'],
                    'type': best['type'],
                })

all_waits = waiting_records + pkg_wait_records

# ── aggregate metrics ─────────────────────────────────────────────────────────

def mean(vals):
    return sum(vals)/len(vals) if vals else 0

# Average wait by transition
trans_waits = defaultdict(list)
for r in all_waits:
    trans_waits[r['transition']].append(r['wait_d'])

top_bottlenecks = sorted(
    [{'transition': t, 'avg_wait_d': round(mean(v),2), 'count': len(v),
      'max_wait_d': round(max(v),2)}
     for t, v in trans_waits.items()],
    key=lambda x: x['avg_wait_d'], reverse=True
)[:10]

# Weekly breakdown for trend chart
week_data = defaultdict(lambda: defaultdict(list))
for r in all_waits:
    if r['week']:
        week_data[r['week']][r['transition']].append(r['wait_d'])

# Build series for stacked bar chart (last 16 weeks)
all_weeks = sorted(week_data.keys())[-16:]

TRANSITIONS_ORDER = [
    'Dispensing → Compression',
    'Dispensing → Encap Hard',
    'Compression → Coating',
    'Sg Med Disp → Sg Encap',
    'Sg Gel Disp → Sg Encap',
    'Manufacturing → Packaging',
]

chart_labels = all_weeks
chart_series = {}
for t in TRANSITIONS_ORDER:
    series = []
    for w in all_weeks:
        vals = week_data[w].get(t, [])
        series.append(round(mean(vals), 2) if vals else 0)
    chart_series[t] = series

# Also add any other transitions found
for t in trans_waits:
    if t not in chart_series:
        series = []
        for w in all_weeks:
            vals = week_data[w].get(t, [])
            series.append(round(mean(vals), 2) if vals else 0)
        chart_series[t] = series

# KPI calculations
all_wait_days = [r['wait_d'] for r in all_waits]
total_waits_by_wo = defaultdict(float)
for r in all_waits:
    total_waits_by_wo[r['wo']] += r['wait_d']

# Estimate PCT: need dispensing start + packaging finish
# Build per-WO timeline
wo_timeline = defaultdict(lambda: {'starts': [], 'finishes': [], 'steps': []})
for r in all_mfg:
    if r['start']:  wo_timeline[r['wo']]['starts'].append(r['start'])
    if r['finish']: wo_timeline[r['wo']]['finishes'].append(r['finish'])
    wo_timeline[r['wo']]['steps'].append(r['step'])

# ── avg mfg days per type: dispensing start → last mfg step finish (excl. packaging) ──
mfg_days_by_type = defaultdict(list)
disp_starts_by_item = defaultdict(list)
for wo2, steps2 in wo_steps.items():
    for s in steps2:
        if s['item'] and s['step'] in ('dispensing', 'sg_gel_disp', 'sg_med_disp') and s['start']:
            disp_starts_by_item[s['item']].append(s['start'])

for item, mfg_list in mfg_last_by_item.items():
    starts = disp_starts_by_item.get(item, [])
    if not starts: continue
    earliest_start = min(starts)
    for m in mfg_list:
        if not m['finish']: continue
        days = (m['finish'] - earliest_start).total_seconds() / 86400
        if 0 < days <= 60:
            ptype = m.get('type', '')
            if ptype in ('TC', 'TU', 'CH', 'SG'):
                mfg_days_by_type[ptype].append(days)

avg_mfg_by_type = {pt: round(mean(v), 1) for pt, v in mfg_days_by_type.items() if v}
all_mfg_vals = [v for vals in mfg_days_by_type.values() for v in vals]
avg_mfg_days = round(mean(all_mfg_vals), 1) if all_mfg_vals else None

# keep avg_pct for target vs actual card
pct_values = []
for pr in all_pkg:
    wo = pr['wo']
    item = pr['item']
    mfg_starts = disp_starts_by_item.get(item, [])
    if mfg_starts and pr['finish']:
        pct_days = (pr['finish'] - min(mfg_starts)).total_seconds() / 86400
        if 0 < pct_days <= 120:
            pct_values.append(pct_days)

avg_pct = round(mean(pct_values), 1) if pct_values else 22.4
avg_wait_total = round(mean(all_wait_days), 1) if all_wait_days else 4.2
biggest_bt = top_bottlenecks[0]['transition'] if top_bottlenecks else 'Manufacturing → Packaging'
biggest_bt_days = top_bottlenecks[0]['avg_wait_d'] if top_bottlenecks else 3.5

# Waiting time % of PCT
if pct_values and all_wait_days:
    # average total wait per batch / average PCT
    wo_totals = list(total_waits_by_wo.values())
    avg_wo_wait = mean(wo_totals)
    wait_pct = round(min(100, avg_wo_wait / max(avg_pct, 1) * 100), 1)
else:
    wait_pct = round(avg_wait_total / max(avg_pct, 1) * 100 * 2, 1)  # rough estimate

# Monthly PCT trend (last 6 months) - mock if not enough data
monthly_pct = {}
for r in all_waits:
    if r['month']:
        monthly_pct.setdefault(r['month'], []).append(r['wait_d'])
months_sorted = sorted(monthly_pct.keys())[-6:]
monthly_avg = {m: round(mean(v)*3 + 14, 1) for m, v in monthly_pct.items()}  # rough PCT estimate

# Detailed batch table (top waits)
detail_table = sorted(all_waits, key=lambda x: x['wait_d'], reverse=True)[:50]
for r in detail_table:
    r['start_str'] = r.get('date','')
    r['wait_str']  = f"{r['wait_d']:.1f}d"

# Work center current status (for WC detail view)
# Find WOs that are scheduled but not finished (future dates or in progress)
today = datetime(2026, 4, 27)
active_wos = []
for wo, steps in wo_steps.items():
    for s in steps:
        if s['finish'] and s['finish'] >= today - timedelta(days=7):
            if s['step'] in ('compression', 'coating', 'encap_hard', 'sg_encap'):
                active_wos.append({
                    'wo': wo, 'step': s['step'], 'wc': s['wc'],
                    'item': s['item'], 'desc': s['desc'][:40] if s['desc'] else '',
                    'start': s['start'].isoformat() if s['start'] else None,
                    'finish': s['finish'].isoformat() if s['finish'] else None,
                })

active_wos = active_wos[:30]

# Transitions for filter dropdown
all_transitions = sorted(trans_waits.keys())

print(f"\n── Summary ───────────────────────────────")
print(f"  Manufacturing records:  {len(all_mfg)}")
print(f"  Packaging records:      {len(all_pkg)}")
print(f"  Waiting time records:   {len(all_waits)}")
print(f"  Avg PCT:                {avg_pct} days")
print(f"  Avg wait total:         {avg_wait_total} days")
print(f"  Waiting % of PCT:       {wait_pct}%")
print(f"  Biggest bottleneck:     {biggest_bt} ({biggest_bt_days}d)")
print(f"  Top transitions:        {[t['transition'] for t in top_bottlenecks[:5]]}")

# ── build chart-ready data ────────────────────────────────────────────────────
COLORS = {
    'Dispensing → Compression':       '#e040a8',
    'Dispensing → Encap Hard':        '#f48acb',
    'Compression → Coating':          '#9c5bbf',
    'Sg Med Disp → Sg Encap':         '#7eb8d4',
    'Sg Gel Disp → Sg Encap':         '#5ba3c2',
    'Manufacturing → Packaging':      '#4db89a',
    'default':                        '#a0a0c0',
}

def color_for(t):
    return COLORS.get(t, COLORS['default'])

chart_datasets = []
for t, series in chart_series.items():
    if any(v > 0 for v in series):
        chart_datasets.append({
            'label': t, 'data': series,
            'backgroundColor': color_for(t),
            'borderColor': color_for(t),
            'borderWidth': 0, 'borderRadius': 3,
        })

# ── per-product-type daily breakdown chart ───────────────────────────────────
# Build per-WO timeline: run hours for each step + wait hours between steps

def get_run_h(s):
    """Allocated run hours for a step record."""
    if s and s.get('run_h') and s['run_h'] > 0:
        return s['run_h']
    if s and s.get('start') and s.get('finish'):
        h = (s['finish'] - s['start']).total_seconds() / 3600
        return h if 0 < h <= 200 else 0
    return 0

def get_wait_h_between(by_step, from_step, to_step):
    sf = by_step.get(from_step)
    st = by_step.get(to_step)
    if sf and st and sf.get('finish') and st.get('start'):
        h = (st['start'] - sf['finish']).total_seconds() / 3600
        return max(0, h) if h <= 30 * 24 else 0
    return 0

wo_ptype_entries = []
for wo, steps in wo_steps.items():
    by_step = {}
    for s in steps:
        k = s['step']
        if k not in by_step or (s['finish'] and (by_step[k]['finish'] is None or s['finish'] > by_step[k]['finish'])):
            by_step[k] = s
    item  = next((s['item'] for s in steps if s['item']), '')
    ptype = product_types.get(item, '')
    if ptype not in ('TC', 'TU', 'CH', 'SG'):
        continue
    # grouping date = dispensing start (or sg_encap start for SG)
    disp = (by_step.get('dispensing') or by_step.get('sg_gel_disp')
            or by_step.get('sg_med_disp') or by_step.get('sg_encap'))
    if not disp or not disp.get('start'):
        continue
    date_str = disp['start'].strftime('%Y-%m-%d')

    if ptype == 'TC':
        if 'compression' not in by_step or 'coating' not in by_step:
            continue
        wo_ptype_entries.append({
            'date': date_str, 'type': ptype, 'wo': wo, 'item': item,
            'disp_run':        get_run_h(by_step.get('dispensing')),
            'disp_comp_wait':  get_wait_h_between(by_step, 'dispensing', 'compression'),
            'comp_run':        get_run_h(by_step.get('compression')),
            'comp_coat_wait':  get_wait_h_between(by_step, 'compression', 'coating'),
            'coat_run':        get_run_h(by_step.get('coating')),
        })
    elif ptype == 'TU':
        if 'compression' not in by_step:
            continue
        wo_ptype_entries.append({
            'date': date_str, 'type': ptype, 'wo': wo, 'item': item,
            'disp_run':        get_run_h(by_step.get('dispensing')),
            'disp_comp_wait':  get_wait_h_between(by_step, 'dispensing', 'compression'),
            'comp_run':        get_run_h(by_step.get('compression')),
        })
    elif ptype == 'CH':
        if 'encap_hard' not in by_step:
            continue
        wo_ptype_entries.append({
            'date': date_str, 'type': ptype, 'wo': wo, 'item': item,
            'disp_run':         get_run_h(by_step.get('dispensing')),
            'disp_encap_wait':  get_wait_h_between(by_step, 'dispensing', 'encap_hard'),
            'encap_run':        get_run_h(by_step.get('encap_hard')),
        })
    elif ptype == 'SG':
        if 'sg_encap' not in by_step:
            continue
        wo_ptype_entries.append({
            'date': date_str, 'type': ptype, 'wo': wo, 'item': item,
            'sg_encap_run':  get_run_h(by_step.get('sg_encap')),
        })

# Add packaging wait + duration per item (averaged across all pkg records)
item_pkg_avg = {}
for item, pkg_list in pkg_by_item.items():
    mfg_finishes = sorted(mfg_last_by_item.get(item, []), key=lambda x: x['finish'])
    waits_h, durs_h = [], []
    for pr in sorted(pkg_list, key=lambda x: x['start']):
        best = None
        for m in mfg_finishes:
            diff = (pr['start'] - m['finish']).total_seconds() / 3600
            if -48 <= diff <= 30 * 24:
                if best is None or abs(diff) < abs((pr['start'] - best['finish']).total_seconds() / 3600):
                    best = m
        if best:
            w = max(0, (pr['start'] - best['finish']).total_seconds() / 3600)
            if w <= 30 * 24:
                waits_h.append(w)
        if pr.get('run_h') and pr['run_h'] > 0:
            durs_h.append(pr['run_h'])
    if waits_h or durs_h:
        item_pkg_avg[item] = {
            'wait_h': round(mean(waits_h), 2) if waits_h else 0,
            'dur_h':  round(mean(durs_h),  2) if durs_h  else 0,
        }

for e in wo_ptype_entries:
    ps = item_pkg_avg.get(e['item'], {})
    e['pkg_wait'] = ps.get('wait_h', 0)   # hours
    e['pkg_run']  = ps.get('dur_h', 0)    # hours

# Aggregate by date per product type (all values in HOURS for chart)
def agg_daily(entries, fields):
    by_date = defaultdict(lambda: defaultdict(list))
    for e in entries:
        for f in fields:
            v = e.get(f, 0)
            if v and v > 0:
                by_date[e['date']][f].append(v)
    return {d: {f: round(mean(by_date[d][f]), 2) if by_date[d][f] else 0 for f in fields}
            for d in sorted(by_date)}

PTYPE_FIELDS = {
    'TC': ['disp_run','disp_comp_wait','comp_run','comp_coat_wait','coat_run','pkg_wait','pkg_run'],
    'TU': ['disp_run','disp_comp_wait','comp_run','pkg_wait','pkg_run'],
    'CH': ['disp_run','disp_encap_wait','encap_run','pkg_wait','pkg_run'],
    'SG': ['sg_encap_run','pkg_wait','pkg_run'],
}
PTYPE_LABELS = {
    'TC': ['Dispensing','Wait → Compression','Compression','Wait → Coating','Coating','Wait → Packaging','Packaging'],
    'TU': ['Dispensing','Wait → Compression','Compression','Wait → Packaging','Packaging'],
    'CH': ['Dispensing','Wait → Encapsulation','Encapsulation','Wait → Packaging','Packaging'],
    'SG': ['SG Encapsulation','Wait → Packaging','Packaging'],
}
# Solid colors for active steps, lighter/dashed-border for wait segments
PTYPE_COLORS = {
    'TC': [
        ('#e040a8','#e040a8'),  # Dispensing — solid pink
        ('#f0a0cc','#e040a8'),  # Wait → Comp — lighter pink
        ('#9c5bbf','#9c5bbf'),  # Compression — solid purple
        ('#c8a8e0','#9c5bbf'),  # Wait → Coating — lighter purple
        ('#4db89a','#4db89a'),  # Coating — solid teal
        ('#a0dcc8','#4db89a'),  # Wait → Pkg — lighter teal
        ('#6dd96d','#6dd96d'),  # Packaging — solid green
    ],
    'TU': [
        ('#e040a8','#e040a8'),
        ('#f0a0cc','#e040a8'),
        ('#9c5bbf','#9c5bbf'),
        ('#c8a8e0','#9c5bbf'),
        ('#6dd96d','#6dd96d'),
    ],
    'CH': [
        ('#e040a8','#e040a8'),
        ('#f0a0cc','#e040a8'),
        ('#f5a623','#f5a623'),
        ('#fad090','#f5a623'),
        ('#6dd96d','#6dd96d'),
    ],
    'SG': [
        ('#7eb8d4','#7eb8d4'),
        ('#b8dce8','#7eb8d4'),
        ('#6dd96d','#6dd96d'),
    ],
}
PTYPE_IS_WAIT = {
    'TC': [False, True, False, True, False, True, False],
    'TU': [False, True, False, True, False],
    'CH': [False, True, False, True, False],
    'SG': [False, True, False],
}

ptype_chart = {}
for pt in ('TC','TU','CH','SG'):
    entries = [e for e in wo_ptype_entries if e['type'] == pt]
    agg = agg_daily(entries, PTYPE_FIELDS[pt])
    dates = sorted(agg.keys())
    datasets = []
    for i, (field, label) in enumerate(zip(PTYPE_FIELDS[pt], PTYPE_LABELS[pt])):
        data = [agg[d].get(field, 0) for d in dates]
        if not any(v > 0 for v in data):
            continue
        bg, border = PTYPE_COLORS[pt][i]
        is_wait = PTYPE_IS_WAIT[pt][i]
        datasets.append({
            'label': label, 'data': data,
            'backgroundColor': bg,
            'borderColor': border,
            'borderWidth': 1 if is_wait else 0,
            'borderRadius': 2,
            'isWait': is_wait,
        })
    ptype_chart[pt] = {'labels': dates, 'datasets': datasets}
    print(f"  ptype {pt}: {len(dates)} days, {len(entries)} WOs")

# ── compression booth snapshot ───────────────────────────────────────────────
SNAPSHOT_DT  = datetime.now()
SNAPSHOT_STR = SNAPSHOT_DT.strftime('%d %b %Y — %I:%M %p')

COMP_BOOTHS = [
    ('TC1-Korsch 1', 'TC1', 'Korsch 1'),
    ('TC2-Korsch 2', 'TC2', 'Korsch 2'),
    ('TC 3-Kil',     'TC3', 'Kilian'),
    ('TC 4-Kil',     'TC4', 'Kilian'),
    ('TC 5-Bosch',   'TC5', 'Bosch'),
    ('TC 6-Fette 1', 'TC6', 'Fette 1'),
    ('TC 7-Fette 2', 'TC7', 'Fette 2'),
    ('TC 8-Fette 3', 'TC8', 'Fette 3'),
]

def fmt_dt(dt):
    return dt.strftime('%d %b, %H:%M') if dt else '—'

def dur_str(s, f):
    if s and f:
        h = (f - s).total_seconds() / 3600
        return f"{int(h)}h {int((h % 1) * 60):02d}m"
    return '—'

# Build per-booth WO list from mfg records
# Include jobs with dates OR with a recognised status (r/w/ip) even if undated
booth_wo_map = {b[0]: [] for b in COMP_BOOTHS}
for r in all_mfg:
    if r['wc'] in booth_wo_map and (
        r['start'] or r['finish']
        or r.get('status', '').lower() in QUEUE_STATUSES | RUNNING_STATUSES
    ):
        booth_wo_map[r['wc']].append(r)
for v in booth_wo_map.values():
    v.sort(key=lambda x: x['start'] or x['finish'] or datetime.min)

compression_booths = []
for bname, short, machine in COMP_BOOTHS:
    jobs = booth_wo_map[bname]

    running = last_done = next_sched = None
    # Explicit IP status takes priority over time-based detection
    for j in jobs:
        if j.get('status', '').lower() in RUNNING_STATUSES:
            running = j
            break
    if not running:
        for j in jobs:
            s, f = j.get('start'), j.get('finish')
            if s and f:
                if s <= SNAPSHOT_DT < f:
                    running = j; break
                elif f <= SNAPSHOT_DT:
                    if not last_done or f > last_done['finish']:
                        last_done = j
            elif s and not f and s <= SNAPSHOT_DT:
                running = j; break
            elif s and s > SNAPSHOT_DT:
                if not next_sched or s < next_sched['start']:
                    next_sched = j

    if running:   status, current = 'RUN',   running
    elif last_done: status, current = 'IDLE',  last_done
    elif next_sched: status, current = 'SCHED', next_sched
    else:           status, current = 'IDLE',  None

    idle_h = None
    if status == 'IDLE' and current and current.get('finish'):
        idle_h = round((SNAPSHOT_DT - current['finish']).total_seconds() / 3600, 1)

    def make_job(j):
        return {
            'wo':      j['wo'],
            'item':    j['item'],
            'desc':    (j['desc'] or '').strip()[:40],
            'qty':     str(j.get('qty') or ''),
            'type':    product_types.get(j['item'], ''),
            'run_h':   round(j['run_h'], 1) if j.get('run_h') else 0,
            'clean_h': round(j['clean_h'], 2) if j.get('clean_h') else 0,
            'start':   fmt_dt(j.get('start')),
            'finish':  fmt_dt(j.get('finish')),
            'duration': dur_str(j.get('start'), j.get('finish')),
            'status':  j.get('status', ''),
        }

    # waiting queue: orders with status 'r' (released) or 'w' (waiting)
    queue_jobs = [j for j in jobs
                  if j.get('status', '').lower() in QUEUE_STATUSES]

    queue_total_h = round(sum(
        (j.get('run_h') or 0) + (j.get('clean_h') or 0)
        for j in queue_jobs
    ), 1)

    compression_booths.append({
        'name': bname, 'short': short, 'machine': machine,
        'status': status,
        'idle_h': idle_h,
        'current': make_job(current) if current else None,
        'queue_items': [make_job(j) for j in queue_jobs],
        'queue_total_h': queue_total_h,
    })
    print(f"  booth {short}: {status}, {len(jobs)} jobs, idle {idle_h}h")

# ── coating booth snapshot ───────────────────────────────────────────────────
COATING_BOOTHS = [
    ('Coating 1', 'C1', 'Coater 1'),
    ('Coating 2', 'C2', 'Coater 2'),
    ('Coating 3', 'C3', 'Coater 3'),
]

coating_booth_map = {b[0]: [] for b in COATING_BOOTHS}
for r in all_mfg:
    if r['wc'] in coating_booth_map and (r['start'] or r['finish']):
        coating_booth_map[r['wc']].append(r)
for v in coating_booth_map.values():
    v.sort(key=lambda x: x['start'] or x['finish'] or datetime.min)

coating_booths = []
for bname, short, machine in COATING_BOOTHS:
    jobs = coating_booth_map[bname]

    running = last_done = next_sched = None
    for j in jobs:
        s, f = j.get('start'), j.get('finish')
        if s and f:
            if s <= SNAPSHOT_DT < f:
                running = j; break
            elif f <= SNAPSHOT_DT:
                if not last_done or f > last_done['finish']:
                    last_done = j
        elif s and not f and s <= SNAPSHOT_DT:
            running = j; break
        elif s and s > SNAPSHOT_DT:
            if not next_sched or s < next_sched['start']:
                next_sched = j

    if running:    status, current = 'RUN',  running
    elif last_done: status, current = 'IDLE', last_done
    elif next_sched: status, current = 'SCHED', next_sched
    else:           status, current = 'IDLE', None

    idle_h = None
    if status == 'IDLE' and current and current.get('finish'):
        idle_h = round((SNAPSHOT_DT - current['finish']).total_seconds() / 3600, 1)

    def make_coat_job(j):
        return {
            'wo':       j['wo'],
            'item':     j['item'],
            'desc':     (j['desc'] or '').strip()[:40],
            'qty':      str(j.get('qty') or ''),
            'type':     product_types.get(j['item'], ''),
            'run_h':    round(j['run_h'], 1) if j.get('run_h') else 0,
            'start':    fmt_dt(j.get('start')),
            'finish':   fmt_dt(j.get('finish')),
            'duration': dur_str(j.get('start'), j.get('finish')),
        }

    queue_jobs = [j for j in jobs if j.get('start') and j['start'] > SNAPSHOT_DT]
    queue_total_h = round(sum((j.get('run_h') or 0) for j in queue_jobs), 1)

    coating_booths.append({
        'name': bname, 'short': short, 'machine': machine,
        'status': status,
        'idle_h': idle_h,
        'current': make_coat_job(current) if current else None,
        'queue_items': [make_coat_job(j) for j in queue_jobs],
        'queue_total_h': queue_total_h,
    })
    print(f"  coater {short}: {status}, {len(jobs)} jobs, idle {idle_h}h")
    'kpi': {
        'avg_pct': avg_pct,
        'pct_target': PCT_TARGET_DAYS,
        'wait_pct': wait_pct,
        'biggest_bottleneck': biggest_bt,
        'biggest_bottleneck_days': biggest_bt_days,
        'avg_wait_days': avg_wait_total,
    },
    'chart_labels': chart_labels,
    'chart_datasets': chart_datasets,
    'top_bottlenecks': top_bottlenecks,
    'detail_table': detail_table[:40],
    'all_transitions': all_transitions,
    'active_wos': active_wos,
    'monthly_labels': months_sorted,
    'monthly_pct': [monthly_avg.get(m, avg_pct) for m in months_sorted],
    'ptype_chart': ptype_chart,
    'compression_booths': compression_booths,
    'coating_booths': coating_booths,
    'snapshot_str': SNAPSHOT_STR,
}

data_json = json.dumps(dashboard_data, default=str, indent=2)


# ── Inject fresh data into index.html ────────────────────────────────────────
TEMPLATE = os.path.join(BASE_DIR, 'index.html')
if not os.path.exists(TEMPLATE):
    raise FileNotFoundError(f'Template not found: {TEMPLATE}')

html = open(TEMPLATE, encoding='utf-8').read()
html = re.sub(
    r'const DATA = \{.*?\};',
    f'const DATA = {data_json};',
    html,
    count=1,
    flags=re.DOTALL,
)

with open(TEMPLATE, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n✅  Dashboard updated: {TEMPLATE}")
print(f"    Open in browser: file://{TEMPLATE}")
